import pandas as pd
import os
import logging
import math
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# --- Configuration ---

# Set up logging to file and console
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("error.log", mode='w'), # Overall log
        logging.StreamHandler()
    ]
)
# Create a separate logger for just errors, as requested by the problem statement
error_logger = logging.getLogger('ErrorLogger')
error_handler = logging.FileHandler('errors.txt', mode='w')
error_handler.setLevel(logging.ERROR)
error_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
error_logger.addHandler(error_handler)

# --- Main Class for Seating Arrangement ---

class SeatingArrangementPlanner:
    """
    Handles the entire process of creating the exam seating arrangement.
    """
    def __init__(self, buffer, arrangement_type, excel_file_path):
        """
        Initializes the planner with user inputs and the single Excel file path.
        """
        self.buffer = buffer
        self.arrangement_type = arrangement_type.lower()
        self.excel_file = excel_file_path

        self.output_dir = "Output"
        os.makedirs(self.output_dir, exist_ok=True)

        self.df_schedule = None
        self.df_enrollment = None
        self.df_names = None
        self.df_rooms = None
        
        self.master_allocation_list = []
        self.master_seats_left_list = []

    def _clean_columns(self, df):
        """Standardizes column names for easier processing."""
        cols = df.columns
        # .lower(), .strip(), and remove dots
        new_cols = [str(c).lower().strip().replace('.', '') for c in cols]
        
        # A robust map to catch variations
        rename_map = {
            'roll': 'roll number', 'rollno': 'roll number',
            'course_code': 'course code', 'course code': 'course code',
            'name': 'name',
            'room no': 'room number', 'room number': 'room number',
            'exam capacity': 'capacity', 'capacity': 'capacity', # Catches 'Exam Capacity' or 'Capacity'
            'block': 'building', 'building': 'building',
            'morning': 'subjects in the morning', 'subjects in the morning': 'subjects in the morning',
            'evening': 'subjects in the evening', 'subjects in the evening': 'subjects in the evening',
            'floor': 'floor' # Keep this in map in case it exists, but we won't use it
        }
        
        final_cols = []
        for c in new_cols:
            mapped = False
            for key, val in rename_map.items():
                if key in c:
                    final_cols.append(val)
                    mapped = True
                    break
            if not mapped:
                final_cols.append(c)
                
        df.columns = final_cols
        return df

    def load_data(self):
        """
        Loads data from all sheets of the single input Excel file.
        """
        logging.info(f"Loading data from single Excel file: {self.excel_file}")
        try:
            # Open the single Excel file
            xls = pd.ExcelFile(self.excel_file)
            
            # Read sheets by their index (more robust than names)
            self.df_schedule = pd.read_excel(xls, sheet_name=0, header=0) # Timetable
            self.df_enrollment = pd.read_excel(xls, sheet_name=1, header=0) # Course-Roll
            self.df_names = pd.read_excel(xls, sheet_name=2, header=0) # Roll-Name
            self.df_rooms = pd.read_excel(xls, sheet_name=3, header=0) # Room Capacity

            # Clean column names for consistency
            self.df_schedule = self._clean_columns(self.df_schedule)
            self.df_enrollment = self._clean_columns(self.df_enrollment)
            self.df_names = self._clean_columns(self.df_names)
            self.df_rooms = self._clean_columns(self.df_rooms)
            
            # Ensure Date column is in datetime format
            self.df_schedule['date'] = pd.to_datetime(self.df_schedule['date'])
            
            # Drop empty/unnamed columns that might be read
            self.df_rooms = self.df_rooms.loc[:, ~self.df_rooms.columns.str.contains('^unnamed')]
            self.df_rooms = self.df_rooms.dropna(how='all') # Drop rows that are completely empty
            
            # Ensure critical columns have correct type
            self.df_rooms['capacity'] = pd.to_numeric(self.df_rooms['capacity'], errors='coerce')
            self.df_rooms = self.df_rooms.dropna(subset=['capacity']) # Remove rooms without capacity
            self.df_rooms['capacity'] = self.df_rooms['capacity'].astype(int)

            # --- 'floor' column creation REMOVED ---

            logging.info("Data loaded and cleaned successfully from all sheets.")
            return True
        except FileNotFoundError:
            logging.error(f"The input file was not found: {self.excel_file}")
            error_logger.error(f"FileNotFoundError: {self.excel_file}")
            return False
        except KeyError as e:
            logging.error(f"Failed to find a required column after cleaning. Missing column: {e}")
            error_logger.error(f"Data loading KeyError: {e}. Check input file headers.")
            return False
        except Exception as e:
            logging.error(f"Failed to read Excel file. Ensure it has 4 sheets. Error: {e}")
            error_logger.error(f"Data loading error: {e}")
            return False

    def get_subject_data_and_check_clashes(self, subject_codes):
        """
        Gathers student data for a list of subjects and checks for registration clashes.
        """
        subjects_data = {}
        logging.info(f"Processing subjects: {', '.join(subject_codes)}")
        for code in subject_codes:
            # Get roll numbers, ensuring they are treated as strings
            try:
                rolls = self.df_enrollment[self.df_enrollment['course code'] == code]['roll number'].astype(str).tolist()
                if not rolls:
                    logging.warning(f"Subject '{code}' has no students enrolled or is not in the enrollment sheet.")
                    continue
                
                sorted_rolls = sorted(rolls)
                subjects_data[code] = {'roll_numbers': sorted_rolls, 'count': len(sorted_rolls)}
            except KeyError:
                msg = f"'course code' or 'roll number' column not found in enrollment sheet after cleaning."
                logging.error(msg)
                error_logger.error(msg)
                return None

        # Clash Detection (as per problem statement)
        roll_counts = defaultdict(list)
        for code in subjects_data:
            for roll in subjects_data[code]['roll_numbers']:
                roll_counts[roll].append(code)

        has_clash = False
        for roll, courses in roll_counts.items():
            if len(courses) > 1:
                clash_msg = f"CLASH DETECTED! Roll number '{roll}' is in multiple courses: {', '.join(courses)}"
                logging.error(clash_msg)
                error_logger.error(clash_msg)
                has_clash = True
        
        if has_clash: return None
            
        logging.info("No clashes found for this session.")
        return subjects_data

    def allocate_students(self, subjects_data):
        """
        Core algorithm to allocate students to rooms based on user settings.
        """
        try:
            session_rooms = self.df_rooms.copy()
            session_rooms['effective_capacity'] = session_rooms['capacity'] - self.buffer
            session_rooms.loc[session_rooms['effective_capacity'] < 0, 'effective_capacity'] = 0
            session_rooms['remaining_capacity'] = session_rooms['effective_capacity']
            session_rooms['allocated_subjects'] = [{} for _ in range(len(session_rooms))]
            
            # Largest course first
            sorted_subject_codes = sorted(subjects_data.keys(), key=lambda s: subjects_data[s]['count'], reverse=True)
            
            # --- THIS IS THE FIX ---
            # Sort rooms to optimize: Group by Building, then largest capacity
            # 'floor' has been removed from the sorting
            session_rooms = session_rooms.sort_values(
                by=['building', 'effective_capacity'], 
                ascending=[True, False]
            ).reset_index(drop=True)
            
            session_allocations = []
            total_students_to_allocate = sum(data['count'] for data in subjects_data.values())
            total_room_capacity = session_rooms['effective_capacity'].sum()

            if total_students_to_allocate > total_room_capacity:
                msg = f"Cannot allocate due to excess students. Required: {total_students_to_allocate}, Available: {total_room_capacity}"
                logging.error(msg)
                error_logger.error(msg)
                return None, None

            for subject_code in sorted_subject_codes:
                students_to_allocate = list(subjects_data[subject_code]['roll_numbers'])
                allocated_in_building = None
                
                while students_to_allocate:
                    available_rooms = session_rooms[session_rooms['remaining_capacity'] > 0]
                    
                    if allocated_in_building:
                        preferred_rooms = available_rooms[available_rooms['building'] == allocated_in_building]
                        target_rooms = preferred_rooms if not preferred_rooms.empty else available_rooms
                    else:
                        target_rooms = available_rooms
                    
                    if target_rooms.empty:
                        msg = f"Ran out of rooms while allocating for {subject_code}. {len(students_to_allocate)} students left."
                        logging.error(msg)
                        error_logger.error(msg)
                        break
                    
                    room_index = target_rooms.index[0] 
                    room_details = session_rooms.loc[room_index]
                    
                    if not allocated_in_building:
                        allocated_in_building = room_details['building']

                    if self.arrangement_type == 'sparse':
                        max_per_subject = math.floor(room_details['effective_capacity'] / 2)
                        already_in_room = room_details['allocated_subjects'].get(subject_code, 0)
                        capacity_for_slot = max(0, min(room_details['remaining_capacity'], max_per_subject - already_in_room))
                    else: # dense
                        capacity_for_slot = room_details['remaining_capacity']

                    if capacity_for_slot <= 0 and len(students_to_allocate) > 0:
                        session_rooms.loc[room_index, 'remaining_capacity'] = -999 
                        continue

                    num_to_place = min(len(students_to_allocate), int(capacity_for_slot))
                    
                    placed_students = students_to_allocate[:num_to_place]
                    students_to_allocate = students_to_allocate[num_to_place:]
                    
                    session_rooms.loc[room_index, 'remaining_capacity'] -= num_to_place
                    current_allocs = session_rooms.at[room_index, 'allocated_subjects']
                    current_allocs[subject_code] = current_allocs.get(subject_code, 0) + num_to_place
                    
                    # 'Floor' removed from this dictionary
                    session_allocations.append({
                        'course_code': subject_code,
                        'Room': str(room_details['room number']),
                        'Building': room_details['building'],
                        'Allocated_students_count': num_to_place,
                        'Roll_list': placed_students
                    })
            
            return session_allocations, session_rooms
        except Exception as e:
            logging.error(f"Error during allocation: {e}")
            error_logger.error(f"Allocation Error: {e}", exc_info=True) # Added exc_info for more detail
            return None, None

    def generate_outputs(self, date_info, session_name, allocations, final_room_state):
        """Generates all output files for a single session"""
        try:
            date_str = date_info.strftime('%Y-%m-%d')
            day_str = date_info.strftime('%A')
            
            session_folder = os.path.join(self.output_dir, f"{date_str}_{day_str}", session_name)
            os.makedirs(session_folder, exist_ok=True)
            
            logging.info(f"Generating output files in: {session_folder}")

            for alloc in allocations:
                # 'Floor' removed from this dictionary
                self.master_allocation_list.append({
                    'Date': date_str, 'Day': day_str, 'Session': session_name, 
                    'Building': alloc['Building'], **alloc
                })
                
                filename = f"{date_str}{alloc['course_code']}{alloc['Room']}.xlsx"
                filepath = os.path.join(session_folder, filename)
                self.create_attendance_sheet(
                    path=filepath, date_str=date_str, session_name=session_name,
                    subject_code=alloc['course_code'], room_number=alloc['Room'],
                    roll_numbers=alloc['Roll_list']
                )

            for _, room in final_room_state.iterrows():
                total_allotted = room['effective_capacity'] - room['remaining_capacity'] if room['remaining_capacity'] > -990 else room['effective_capacity']
                self.master_seats_left_list.append({
                    'Date': date_str, 'Session': session_name,
                    'Room No.': str(room['room number']), 'Block': room['building'],
                    'Exam Capacity': room['capacity'], 'Alloted': total_allotted,
                    'Vacant (B-C)': room['capacity'] - total_allotted
                })
        except Exception as e:
            logging.error(f"Error generating outputs: {e}")
            error_logger.error(f"Output Generation Error: {e}")

    def create_attendance_sheet(self, path, date_str, session_name, subject_code, room_number, roll_numbers):
        """Creates the formatted daily attendance sheet for one room"""
        try:
            self.df_names['roll number'] = self.df_names['roll number'].astype(str)
            rolls_df = pd.DataFrame(roll_numbers, columns=['roll number'])
            
            student_details = pd.merge(rolls_df, self.df_names, on='roll number', how='left')
            student_details['name'].fillna('Unknown Name', inplace=True) 
            student_details = student_details[['roll number', 'name']]
            student_details.columns = ['Roll Number', 'Student Name']
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"

            header_font = Font(bold=True, size=14)
            col_header_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            header_text = f"Course: {subject_code} | Room: {room_number} | Date: {date_str} | Session: {session_name}"
            ws.merge_cells('A1:B1')
            cell = ws['A1']
            cell.value = header_text
            cell.font = header_font
            cell.alignment = center_align
            ws.row_dimensions[1].height = 30

            start_row = 3
            for c_idx, col_name in enumerate(student_details.columns, 1):
                cell = ws.cell(row=start_row, column=c_idx, value=col_name)
                cell.font = col_header_font
            
            for r_idx, row in enumerate(dataframe_to_rows(student_details, index=False, header=False), start_row + 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            last_student_row = ws.max_row
            
            static_footer = [("Total Students:", len(roll_numbers))] + \
                            [(f"TA {i}", "") for i in range(1, 6)] + \
                            [(f"Invigilator {i}", "") for i in range(1, 6)]
            
            for i, (label, value) in enumerate(static_footer, 1):
                row_num = last_student_row + 2 + i
                ws.cell(row=row_num, column=1, value=label).font = Font(bold=True if 'Students' in label else False)
                ws.cell(row=row_num, column=2, value=value)

            ws.column_dimensions[get_column_letter(1)].width = 25
            ws.column_dimensions[get_column_letter(2)].width = 40
                
            wb.save(path)
        except Exception as e:
            logging.error(f"Could not save attendance sheet {path}. Error: {e}")
            error_logger.error(f"Failed to save {path}: {e}")

    def finalize_reports(self):
        """Creates the two final summary Excel files"""
        logging.info("Generating final summary reports...")
        
        try:
            # --- Report 1: Overall Seating Arrangement ---
            if self.master_allocation_list:
                df_overall = pd.DataFrame(self.master_allocation_list)
                df_overall['Roll_list (semicolon separated_,'] = df_overall['Roll_list'].apply(lambda r: ';'.join(map(str, r)))
                
                final_cols = ['Date', 'Day', 'course_code', 'Room', 'Allocated_students_count', 'Roll_list (semicolon separated_,']
                df_overall = df_overall[final_cols]
                
                filepath = os.path.join(self.output_dir, "op_overall_seating_arrangement.xlsx")
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    df_overall.to_excel(writer, sheet_name='op_seating_arrangement', index=False, startrow=1)
                    ws = writer.sheets['op_seating_arrangement']
                    ws['A1'] = "Seating Plan"
                    ws['A1'].font = Font(bold=True)

                logging.info(f"Successfully created '{filepath}'")

            # --- Report 2: Seats Left ---
            if self.master_seats_left_list:
                df_seats_left = pd.DataFrame(self.master_seats_left_list)
                df_seats_left = df_seats_left.drop_duplicates() 
                
                filepath = os.path.join(self.output_dir, "op_seats_left.xlsx")
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    for (date, session), group in df_seats_left.groupby(['Date', 'Session']):
                        sheet_name = f"{date}_{session}"
                        
                        final_cols = ['Room No.', 'Exam Capacity', 'Block', 'Alloted', 'Vacant (B-C)']
                        group_to_write = group[final_cols]
                        
                        group_to_write.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                        
                        ws = writer.sheets[sheet_name]
                        ws.merge_cells('A1:E1') # 5 columns
                        ws['A1'] = f"Date: {date} - Session: {session}"
                        ws['A1'].font = Font(bold=True)
                        ws['A1'].alignment = Alignment(horizontal='center')

                logging.info(f"Successfully created '{filepath}'")
        except Exception as e:
            logging.error(f"Error finalizing reports: {e}")
            error_logger.error(f"Final Report Error: {e}")

    def run(self):
        """
        Executes the entire seating arrangement process.
        """
        if not self.load_data():
            logging.error("Failed to load data. Exiting.")
            return

        for _, row in self.df_schedule.iterrows():
            date_info = row['date']
            
            for session_name in ['morning', 'evening']:
                col_name = f'subjects in the {session_name}'
                if col_name not in row or pd.isna(row[col_name]) or not str(row[col_name]).strip() or 'no exam' in str(row[col_name]).lower():
                    continue

                logging.info(f"--- Processing Session: {date_info.strftime('%d-%b-%Y')} {session_name.capitalize()} ---")
                
                subjects_str = str(row[col_name])
                subject_codes = [s.strip() for s in subjects_str.split(';') if s.strip()]
                
                subjects_data = self.get_subject_data_and_check_clashes(subject_codes)
                if subjects_data is None:
                    logging.error("Process for this session stopped due to clashes.")
                    continue
                if not subjects_data:
                    logging.warning("No students found for any subject in this session. Skipping.")
                    continue

                allocations, final_room_state = self.allocate_students(subjects_data)
                
                if allocations is not None:
                    self.generate_outputs(date_info, session_name.capitalize(), allocations, final_room_state)
        
        self.finalize_reports()
        logging.info("--- Seating Arrangement Process Completed ---")

def main():
    """
    Main function to get user input and run the planner.
    """
    print("--- Exam Seating Arrangement Generator ---")
    
    # Define the input file name. Assumes it's in the same folder.
    excel_file = 'input_data_tt.xlsx'
    
    try:
        if not os.path.exists(excel_file):
            logging.error(f"FATAL ERROR: Input file '{excel_file}' not found.")
            logging.error("Please make sure 'input_data_tt.xlsx' is in the same folder as this script.")
            error_logger.error(f"FileNotFoundError: {excel_file} not found in script directory.")
            return 

        buffer = int(input("Enter the seat buffer (e.g., 5): "))
        arrangement_type = ""
        while arrangement_type not in ['sparse', 'dense']:
            arrangement_type = input("Enter arrangement type ('sparse' or 'dense'): ").lower()

        planner = SeatingArrangementPlanner(
            buffer=buffer, 
            arrangement_type=arrangement_type,
            excel_file_path=excel_file
        )
        planner.run()

    except ValueError:
        logging.error("Invalid input. Buffer must be an integer. Please restart.")
        error_logger.error("ValueError: Buffer was not an integer.")
    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")
        error_logger.error(f"Unhandled Exception: {e}", exc_info=True)

if __name__ == "__main__":
    main()