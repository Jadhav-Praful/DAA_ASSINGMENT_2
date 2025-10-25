import streamlit as st
import pandas as pd
import os
import logging
import math
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import io
import zipfile

# --- UI Configuration ---
# This line MUST be the first Streamlit command
st.set_page_config(page_title="Exam Seating Planner", layout="wide")
st.title("🎓 Exam Seating Arrangement Generator")

# --- Logging Setup ---
# Use a stringIO buffer to capture log messages to display in Streamlit
log_stream = io.StringIO()

# Configure the root logger
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(log_stream) # Captures all logs
    ]
)
# Configure a separate error logger (can be used for a separate error file if needed)
error_logger = logging.getLogger('ErrorLogger')
error_handler = logging.StreamHandler(log_stream) # Also send errors to the main log display
error_handler.setLevel(logging.ERROR)
error_handler.setFormatter(logging.Formatter('%(asctime)s - ERROR - %(message)s'))
error_logger.addHandler(error_handler)


# --- Main Class for Seating Arrangement ---
class SeatingArrangementPlanner:
    """
    Handles the entire process of creating the exam seating arrangement.
    Modified to work with Streamlit's in-memory file handling.
    """
    def __init__(self, buffer, arrangement_type, uploaded_excel_file):
        """
        Initializes the planner with user inputs from Streamlit.
        """
        self.buffer = buffer
        self.arrangement_type = arrangement_type.lower()
        self.excel_file = uploaded_excel_file # This is now an in-memory file object

        # Store generated files in memory {filename: bytes_io_object}
        self.output_files = {}

        self.df_schedule = None
        self.df_enrollment = None
        self.df_names = None
        self.df_rooms = None
        
        self.master_allocation_list = []
        self.master_seats_left_list = []

    def _clean_columns(self, df):
        """Standardizes column names for easier processing."""
        cols = df.columns
        new_cols = [str(c).lower().strip().replace('.', '') for c in cols]
        
        rename_map = {
            'roll': 'roll number', 'rollno': 'roll number',
            'course_code': 'course code', 'course code': 'course code',
            'name': 'name',
            'room no': 'room number', 'room number': 'room number',
            'exam capacity': 'capacity', 'capacity': 'capacity',
            'block': 'building', 'building': 'building',
            'morning': 'subjects in the morning', 'subjects in the morning': 'subjects in the morning',
            'evening': 'subjects in the evening', 'subjects in the evening': 'subjects in the evening',
            'floor': 'floor'
        }
        
        final_cols = []
        for c in new_cols:
            mapped = False
            for key, val in rename_map.items():
                if key in c:
                    final_cols.append(val)
                    mapped = True
                    break
            if not mapped:
                final_cols.append(c)
                
        df.columns = final_cols
        return df

    def load_data(self):
        """
        Loads data from the uploaded Excel file object.
        """
        logging.info(f"Loading data from uploaded Excel file...")
        try:
            xls = pd.ExcelFile(self.excel_file)
            
            self.df_schedule = pd.read_excel(xls, sheet_name=0, header=0)
            self.df_enrollment = pd.read_excel(xls, sheet_name=1, header=0)
            self.df_names = pd.read_excel(xls, sheet_name=2, header=0)
            self.df_rooms = pd.read_excel(xls, sheet_name=3, header=0)

            self.df_schedule = self._clean_columns(self.df_schedule)
            self.df_enrollment = self._clean_columns(self.df_enrollment)
            self.df_names = self._clean_columns(self.df_names)
            self.df_rooms = self._clean_columns(self.df_rooms)
            
            self.df_schedule['date'] = pd.to_datetime(self.df_schedule['date'])
            
            self.df_rooms = self.df_rooms.loc[:, ~self.df_rooms.columns.str.contains('^unnamed')]
            self.df_rooms = self.df_rooms.dropna(how='all')
            
            self.df_rooms['capacity'] = pd.to_numeric(self.df_rooms['capacity'], errors='coerce')
            self.df_rooms = self.df_rooms.dropna(subset=['capacity'])
            self.df_rooms['capacity'] = self.df_rooms['capacity'].astype(int)

            logging.info("Data loaded and cleaned successfully from all sheets.")
            return True
        except KeyError as e:
            logging.error(f"Failed to find a required column after cleaning. Missing column: {e}")
            error_logger.error(f"Data loading KeyError: {e}. Check input file headers.")
            return False
        except Exception as e:
            logging.error(f"Failed to read Excel file. Ensure it has 4 sheets in the correct order. Error: {e}")
            error_logger.error(f"Data loading error: {e}")
            return False

    def get_subject_data_and_check_clashes(self, subject_codes):
        """
        Gathers student data for a list of subjects and checks for registration clashes.
        """
        subjects_data = {}
        logging.info(f"Processing subjects: {', '.join(subject_codes)}")
        for code in subject_codes:
            try:
                rolls = self.df_enrollment[self.df_enrollment['course code'] == code]['roll number'].astype(str).tolist()
                if not rolls:
                    logging.warning(f"Subject '{code}' has no students enrolled or is not in the enrollment sheet.")
                    continue
                
                sorted_rolls = sorted(rolls)
                subjects_data[code] = {'roll_numbers': sorted_rolls, 'count': len(sorted_rolls)}
            except KeyError:
                msg = f"'course code' or 'roll number' column not found in enrollment sheet after cleaning."
                logging.error(msg)
                error_logger.error(msg)
                return None

        # Clash Detection
        roll_counts = defaultdict(list)
        for code in subjects_data:
            for roll in subjects_data[code]['roll_numbers']:
                roll_counts[roll].append(code)

        has_clash = False
        for roll, courses in roll_counts.items():
            if len(courses) > 1:
                clash_msg = f"CLASH DETECTED! Roll number '{roll}' is in multiple courses: {', '.join(courses)}"
                logging.error(clash_msg)
                error_logger.error(clash_msg)
                has_clash = True
        
        if has_clash: return None
            
        logging.info("No clashes found for this session.")
        return subjects_data

    def allocate_students(self, subjects_data):
        """
        Core algorithm to allocate students to rooms based on user settings.
        """
        try:
            session_rooms = self.df_rooms.copy()
            session_rooms['effective_capacity'] = session_rooms['capacity'] - self.buffer
            session_rooms.loc[session_rooms['effective_capacity'] < 0, 'effective_capacity'] = 0
            session_rooms['remaining_capacity'] = session_rooms['effective_capacity']
            session_rooms['allocated_subjects'] = [{} for _ in range(len(session_rooms))]
            
            sorted_subject_codes = sorted(subjects_data.keys(), key=lambda s: subjects_data[s]['count'], reverse=True)
            
            session_rooms = session_rooms.sort_values(
                by=['building', 'effective_capacity'], 
                ascending=[True, False]
            ).reset_index(drop=True)
            
            session_allocations = []
            total_students_to_allocate = sum(data['count'] for data in subjects_data.values())
            total_room_capacity = session_rooms['effective_capacity'].sum()

            if total_students_to_allocate > total_room_capacity:
                msg = f"Cannot allocate due to excess students. Required: {total_students_to_allocate}, Available: {total_room_capacity}"
                logging.error(msg)
                error_logger.error(msg)
                return None, None

            for subject_code in sorted_subject_codes:
                students_to_allocate = list(subjects_data[subject_code]['roll_numbers'])
                allocated_in_building = None
                
                while students_to_allocate:
                    available_rooms = session_rooms[session_rooms['remaining_capacity'] > 0]
                    
                    if allocated_in_building:
                        preferred_rooms = available_rooms[available_rooms['building'] == allocated_in_building]
                        target_rooms = preferred_rooms if not preferred_rooms.empty else available_rooms
                    else:
                        target_rooms = available_rooms
                    
                    if target_rooms.empty:
                        msg = f"Ran out of rooms while allocating for {subject_code}. {len(students_to_allocate)} students left."
                        logging.error(msg)
                        error_logger.error(msg)
                        break
                    
                    room_index = target_rooms.index[0] 
                    room_details = session_rooms.loc[room_index]
                    
                    if not allocated_in_building:
                        allocated_in_building = room_details['building']

                    if self.arrangement_type == 'sparse':
                        max_per_subject = math.floor(room_details['effective_capacity'] / 2)
                        already_in_room = room_details['allocated_subjects'].get(subject_code, 0)
                        capacity_for_slot = max(0, min(room_details['remaining_capacity'], max_per_subject - already_in_room))
                    else: # dense
                        capacity_for_slot = room_details['remaining_capacity']

                    if capacity_for_slot <= 0 and len(students_to_allocate) > 0:
                        session_rooms.loc[room_index, 'remaining_capacity'] = -999 
                        continue

                    num_to_place = min(len(students_to_allocate), int(capacity_for_slot))
                    
                    placed_students = students_to_allocate[:num_to_place]
                    students_to_allocate = students_to_allocate[num_to_place:]
                    
                    session_rooms.loc[room_index, 'remaining_capacity'] -= num_to_place
                    current_allocs = session_rooms.at[room_index, 'allocated_subjects']
                    current_allocs[subject_code] = current_allocs.get(subject_code, 0) + num_to_place
                    
                    session_allocations.append({
                        'course_code': subject_code,
                        'Room': str(room_details['room number']),
                        'Building': room_details['building'],
                        'Allocated_students_count': num_to_place,
                        'Roll_list': placed_students
                    })
            
            return session_allocations, session_rooms
        except Exception as e:
            logging.error(f"Error during allocation: {e}")
            error_logger.error(f"Allocation Error: {e}", exc_info=True)
            return None, None

    def generate_outputs(self, date_info, session_name, allocations, final_room_state):
        """Generates all output files for a single session and stores them in memory"""
        try:
            date_str = date_info.strftime('%Y-%m-%d')
            day_str = date_info.strftime('%A')
            
            # Use os.path.join to create platform-agnostic paths for the zip file
            session_folder = os.path.join(f"{date_str}_{day_str}", session_name)
            
            logging.info(f"Generating output files for session: {session_folder}")

            for alloc in allocations:
                self.master_allocation_list.append({
                    'Date': date_str, 'Day': day_str, 'Session': session_name, 
                    'Building': alloc['Building'], **alloc
                })
                
                filename = os.path.join(session_folder, f"{date_str}_{alloc['course_code']}_{alloc['Room']}.xlsx")
                
                # Create attendance sheet in memory
                file_bytes = self.create_attendance_sheet(
                    date_str=date_str, session_name=session_name,
                    subject_code=alloc['course_code'], room_number=alloc['Room'],
                    roll_numbers=alloc['Roll_list']
                )
                if file_bytes:
                    self.output_files[filename] = file_bytes

            for _, room in final_room_state.iterrows():
                total_allotted = room['effective_capacity'] - room['remaining_capacity'] if room['remaining_capacity'] > -990 else room['effective_capacity']
                self.master_seats_left_list.append({
                    'Date': date_str, 'Session': session_name,
                    'Room No.': str(room['room number']), 'Block': room['building'],
                    'Exam Capacity': room['capacity'], 'Alloted': total_allotted,
                    'Vacant (B-C)': room['capacity'] - total_allotted
                })
        except Exception as e:
            logging.error(f"Error generating outputs: {e}")
            error_logger.error(f"Output Generation Error: {e}")

    def create_attendance_sheet(self, date_str, session_name, subject_code, room_number, roll_numbers):
        """Creates the formatted attendance sheet in memory and returns its bytes"""
        try:
            self.df_names['roll number'] = self.df_names['roll number'].astype(str)
            rolls_df = pd.DataFrame(roll_numbers, columns=['roll number'])
            
            student_details = pd.merge(rolls_df, self.df_names, on='roll number', how='left')
            student_details['name'].fillna('Unknown Name', inplace=True) 
            student_details = student_details[['roll number', 'name']]
            student_details.columns = ['Roll Number', 'Student Name']
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"

            header_font = Font(bold=True, size=14)
            col_header_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            header_text = f"Course: {subject_code} | Room: {room_number} | Date: {date_str} | Session: {session_name}"
            ws.merge_cells('A1:B1')
            cell = ws['A1']
            cell.value = header_text
            cell.font = header_font
            cell.alignment = center_align
            ws.row_dimensions[1].height = 30

            start_row = 3
            for c_idx, col_name in enumerate(student_details.columns, 1):
                cell = ws.cell(row=start_row, column=c_idx, value=col_name)
                cell.font = col_header_font
            
            for r_idx, row in enumerate(dataframe_to_rows(student_details, index=False, header=False), start_row + 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            last_student_row = ws.max_row
            
            static_footer = [("Total Students:", len(roll_numbers))] + \
                            [(f"TA {i}", "") for i in range(1, 6)] + \
                            [(f"Invigilator {i}", "") for i in range(1, 6)]
            
            for i, (label, value) in enumerate(static_footer, 1):
                row_num = last_student_row + 2 + i
                ws.cell(row=row_num, column=1, value=label).font = Font(bold=True if 'Students' in label else False)
                ws.cell(row=row_num, column=2, value=value)

            ws.column_dimensions[get_column_letter(1)].width = 25
            ws.column_dimensions[get_column_letter(2)].width = 40
            
            # Save to a bytes buffer instead of a file
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            return output_buffer

        except Exception as e:
            logging.error(f"Could not create attendance sheet for {subject_code} in room {room_number}. Error: {e}")
            error_logger.error(f"Failed to create attendance sheet: {e}")
            return None

    def finalize_reports(self):
        """Creates the two final summary Excel files in memory"""
        logging.info("Generating final summary reports...")
        
        try:
            # --- Report 1: Overall Seating Arrangement ---
            if self.master_allocation_list:
                df_overall = pd.DataFrame(self.master_allocation_list)
                df_overall['Roll_list (semicolon separated)'] = df_overall['Roll_list'].apply(lambda r: ';'.join(map(str, r)))
                
                final_cols = ['Date', 'Day', 'course_code', 'Room', 'Allocated_students_count', 'Roll_list (semicolon separated)']
                df_overall = df_overall[final_cols]
                
                # Save to bytes buffer
                output_buffer = io.BytesIO()
                with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                    df_overall.to_excel(writer, sheet_name='op_seating_arrangement', index=False, startrow=1)
                    ws = writer.sheets['op_seating_arrangement']
                    ws['A1'] = "Seating Plan"
                    ws['A1'].font = Font(bold=True)
                
                output_buffer.seek(0)
                self.output_files["op_overall_seating_arrangement.xlsx"] = output_buffer
                logging.info("Successfully created 'op_overall_seating_arrangement.xlsx'")

            # --- Report 2: Seats Left ---
            if self.master_seats_left_list:
                df_seats_left = pd.DataFrame(self.master_seats_left_list)
                df_seats_left = df_seats_left.drop_duplicates() 
                
                # Save to bytes buffer
                output_buffer = io.BytesIO()
                with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                    for (date, session), group in df_seats_left.groupby(['Date', 'Session']):
                        sheet_name = f"{date}_{session}"
                        final_cols = ['Room No.', 'Exam Capacity', 'Block', 'Alloted', 'Vacant (B-C)']
                        group_to_write = group[final_cols]
                        
                        group_to_write.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                        
                        ws = writer.sheets[sheet_name]
                        ws.merge_cells('A1:E1')
                        ws['A1'] = f"Date: {date} - Session: {session}"
                        ws['A1'].font = Font(bold=True)
                        ws['A1'].alignment = Alignment(horizontal='center')
                
                output_buffer.seek(0)
                self.output_files["op_seats_left.xlsx"] = output_buffer
                logging.info("Successfully created 'op_seats_left.xlsx'")

        except Exception as e:
            logging.error(f"Error finalizing reports: {e}")
            error_logger.error(f"Final Report Error: {e}")

    def run(self):
        """
        Executes the entire seating arrangement process.
        """
        if not self.load_data():
            logging.error("Failed to load data. Exiting.")
            return

        for _, row in self.df_schedule.iterrows():
            date_info = row['date']
            
            for session_name in ['morning', 'evening']:
                col_name = f'subjects in the {session_name}'
                if col_name not in row or pd.isna(row[col_name]) or not str(row[col_name]).strip() or 'no exam' in str(row[col_name]).lower():
                    continue

                logging.info(f"--- Processing Session: {date_info.strftime('%d-%b-%Y')} {session_name.capitalize()} ---")
                
                subjects_str = str(row[col_name])
                subject_codes = [s.strip() for s in subjects_str.split(';') if s.strip()]
                
                subjects_data = self.get_subject_data_and_check_clashes(subject_codes)
                if subjects_data is None:
                    logging.error("Process for this session stopped due to clashes.")
                    continue
                if not subjects_data:
                    logging.warning("No students found for any subject in this session. Skipping.")
                    continue

                allocations, final_room_state = self.allocate_students(subjects_data)
                
                if allocations is not None:
                    self.generate_outputs(date_info, session_name.capitalize(), allocations, final_room_state)
        
        self.finalize_reports()
        logging.info("--- Seating Arrangement Process Completed ---")


# --- Streamlit UI ---

with st.sidebar:
    st.header("⚙️ Settings")
    
    uploaded_file = st.file_uploader(
        "Upload Input Excel File",
        type=["xlsx"],
        help="Upload the Excel file containing the 4 required sheets: Timetable, Course-Roll, Roll-Name, Room-Capacity."
    )
    
    buffer = st.number_input(
        "Enter Seat Buffer", 
        min_value=0, 
        value=5, 
        help="Number of seats to leave empty in each room for spacing."
    )
    
    arrangement_type = st.selectbox(
        "Select Arrangement Type", 
        ('sparse', 'dense'), 
        help="'sparse' attempts to not fill rooms completely, 'dense' fills them up."
    )
    
    start_button = st.button("🚀 Generate Seating Arrangement")

if start_button:
    if uploaded_file is None:
        st.error("❌ Please upload the input Excel file before generating.")
    else:
        with st.spinner('Processing... Please wait.'):
            # Clear previous log
            log_stream.truncate(0)
            log_stream.seek(0)
            
            try:
                planner = SeatingArrangementPlanner(
                    buffer=buffer,
                    arrangement_type=arrangement_type,
                    uploaded_excel_file=uploaded_file
                )
                planner.run()

                # After running, check if there are files to be zipped
                if planner.output_files:
                    st.success("✅ Process completed successfully!")
                    
                    # Create a zip file in memory
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_f:
                        for filename, file_bytes in planner.output_files.items():
                            zip_f.writestr(filename, file_bytes.getvalue())
                    
                    zip_buffer.seek(0)
                    
                    st.download_button(
                        label="⬇️ Download All Output Files (.zip)",
                        data=zip_buffer,
                        file_name=f"Seating_Arrangement_Output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                        mime="application/zip"
                    )
                else:
                    st.warning("⚠️ Process completed, but no output files were generated. Check the logs for details.")

            except Exception as e:
                st.error(f"An unexpected error occurred: {e}")
                logging.error(f"An unhandled exception occurred in the main process: {e}", exc_info=True)

        # Display logs
        st.subheader("Process Logs")
        log_content = log_stream.getvalue()
        st.code(log_content)
else:
    st.info("Please upload your data file and click 'Generate' in the sidebar to begin.")

